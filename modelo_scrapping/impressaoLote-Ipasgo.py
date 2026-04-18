from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
import openpyxl
import logging
import traceback
import os
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
import time

# Caminho para a planilha
workbook_path = r"G:\Meu Drive\2.Faturamento\2 - AUDITORIA\AUTOMACOES-Python\BaseImpressao.xlsx"

# Configuração de logging
LOG_PATH = os.path.join(os.path.dirname(__file__), "log_execucao.txt")
logging.basicConfig(
    filename=LOG_PATH,
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s"
)
_console = logging.StreamHandler()
_console.setLevel(logging.INFO)
_console.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
logging.getLogger().addHandler(_console)

DEFAULT_WAIT_SECONDS = 15

def log_exception(e, contexto: str, locator: str | None = None):
    """Loga exceções com a linha do arquivo onde ocorreu.

    - contexto: descrição do que estávamos tentando fazer (ex.: localizar elemento X)
    - locator: seletor/identificador do elemento (opcional)
    """
    try:
        frames = traceback.extract_tb(e.__traceback__)
        this_name = os.path.basename(__file__)
        lineno = None
        filename = None
        for f in reversed(frames):
            if os.path.basename(f.filename) == this_name:
                lineno = f.lineno
                filename = f.filename
                break
        trace_str = "".join(traceback.format_exception(type(e), e, e.__traceback__))
        logging.error(
            f"Falha em '{contexto}'. Locator='{locator}'. Arquivo='{filename}' Linha={lineno}. Exceção={repr(e)}\n{trace_str}"
        )
    except Exception as log_err:
        # Em caso raro de falha ao formatar o traceback, ainda registrar algo útil
        logging.error(f"Erro ao logar exceção de '{contexto}': {repr(log_err)} | Exceção original: {repr(e)}")

def capture_screenshot(driver, name_hint: str):
    try:
        ts = int(time.time())
        safe = "".join(c for c in name_hint if c.isalnum() or c in ("-", "_"))
        path = os.path.join(os.path.dirname(__file__), f"screenshot_{safe}_{ts}.png")
        driver.save_screenshot(path)
        logging.info(f"Screenshot salvo em {path}")
    except Exception as e:
        logging.warning(f"Falha ao salvar screenshot: {repr(e)}")

def wait_for(driver, by, locator, timeout: int = DEFAULT_WAIT_SECONDS, contexto: str = ""):
    try:
        logging.info(f"Aguardando elemento {by}={locator} em {contexto}")
        elem = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((by, locator)))
        logging.info(f"Elemento presente {by}={locator} em {contexto}")
        return elem
    except Exception as e:
        log_exception(e, f"localizar elemento em {contexto}", locator=f"{by}={locator}")
        try:
            capture_screenshot(driver, f"falha_{by}_{locator}")
        except Exception:
            pass
        raise

def open_prestador():
    # Caminho para o perfil do Chrome
    #user_data_path = r"C:\Users\RECEPCAO_CLMF-001\AppData\Local\Google\Chrome\User Data"
    #profile_name = "Default"  # ou o nome do perfil que você deseja usar

    # Configuração do WebDriver
    options = webdriver.ChromeOptions()
    #options.add_argument(f"user-data-dir={user_data_path}")  # Define o diretório de dados do usuário
    #options.add_argument(f"profile-directory={profile_name}")  # Define o perfil a ser utilizado
    options.add_argument("--kiosk-printing")  # Ativa a impressão direta em modo quiosque
    options.add_argument("--start-maximized")  # Inicia o navegador maximizado

    # Inicializa o driver com as opções configuradas
    driver = webdriver.Chrome(options=options)
    driver.get("https://portalos.ipasgo.go.gov.br/Portal_Dominio/PrestadorLogin.aspx")
    sleep(5)

    try:
        # Login34
        try:
            driver.find_element(By.XPATH, "//*[@id='SilkUIFramework_wt13_block_wtUsername_wtUserNameInput2']").clear()
            driver.find_element(By.XPATH, "//*[@id='SilkUIFramework_wt13_block_wtUsername_wtUserNameInput2']").send_keys("14898500")
        except Exception as e:
            log_exception(e, "preencher usuário", locator="XPATH=SilkUIFramework_wt13...wtUserNameInput2")
            raise
        sleep(1)
        try:
            driver.find_element(By.XPATH, "//*[@id='SilkUIFramework_wt13_block_wtPassword_wtPasswordInput']").clear()
            sleep(1)
            driver.find_element(By.XPATH, "//*[@id='SilkUIFramework_wt13_block_wtPassword_wtPasswordInput']").send_keys("Clmf2025")
        except Exception as e:
            log_exception(e, "preencher senha", locator="XPATH=SilkUIFramework_wt13...wtPasswordInput")
            raise
        try:
            login_btn = wait_for(driver, By.XPATH, "//*[@id='SilkUIFramework_wt13_block_wtAction_wtLoginButton']", contexto="login")
            login_btn.click()
        except Exception as e:
            log_exception(e, "clicar login", locator="By.XPATH=...wtLoginButton")
            raise
        sleep(7)
        

        # Verificar se existe frame/popup para fechar após login
        try:
            logging.info("Verificando se há popup inicial...")
            sleep(3) # Wait robusto após login
            
            popup_xpath = "/html/body/div[1]/div[1]/a"
            if len(driver.find_elements(By.XPATH, popup_xpath)) > 0:
                driver.find_element(By.XPATH, popup_xpath).click()
                logging.info("Popup fechado.")
                sleep(2)
            else:
                 logging.info("Popup não encontrado (sem erro).")
            
            # Aguardar overlay sumir
            try:
                WebDriverWait(driver, 5).until(EC.invisibility_of_element_located((By.CLASS_NAME, "os-internal-ui-widget-overlay")))
            except:
                pass

        except Exception as e:
             logging.info(f"Erro no manuseio do popup: {str(e)}")

        # Navegação para o módulo
        try:
            # Wait explícito para overlay sumir antes de clicar no módulo
            try:
                WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.CLASS_NAME, "os-internal-ui-widget-overlay")))
            except:
                logging.warning("Aviso: Overlay ainda presente ou timeout.")

            mod_xpath = "/html/body/form/div[3]/div/div[2]/div/div[2]/div/div/span/div[5]/div/div[2]/div[2]/div/div/div[2]/table/tbody/tr[2]/td/div/a"
            mod_btn = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, mod_xpath)))
            
            # JS Click
            driver.execute_script("arguments[0].scrollIntoView(true);", mod_btn)
            sleep(1)
            driver.execute_script("arguments[0].click();", mod_btn)
            logging.info("Módulo selecionado (JS click). Aguardando nova aba...")
            
            # Aguardar abertura de nova aba
            for _ in range(20):
                if len(driver.window_handles) > 1:
                    logging.info("Nova aba detectada. Alternando...")
                    driver.switch_to.window(driver.window_handles[-1])
                    break
                sleep(1)
            else:
                logging.warning("Aviso: Nova aba não abriu após 20s.")
        except Exception as e:
            log_exception(e, "acessar módulo após login", locator="/html/body/form/div[3]/div/div[2]/div/div[2]/div/div/span/div[5]/div/div[2]/div[2]/div/div/div[2]/table/tbody/tr[2]/td/div/a")
            raise
        sleep(5)
        driver.switch_to.window(driver.window_handles[1])
        sleep(2)

        # Espera explícita
        wait = WebDriverWait(driver, DEFAULT_WAIT_SECONDS)

        # Navegar até o módulo de faturamento
        LInkLocalizarGuias = "https://novowebplanipasgo.facilinformatica.com.br/GuiasTISS/LocalizarProcedimentos"
        driver.get(LInkLocalizarGuias)
        sleep(5)
        
        # Tentar fechar modal inicial de forma robusta e aguardar overlay
        try:
            logging.info("Verificando modal 'button-1'...")
            
            if len(driver.find_elements(By.ID, "button-1")) > 0:
                btn_close = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, "button-1")))
                btn_close.click()
                logging.info("Modal 'button-1' clicado. Aguardando sumir...")
                
                # Aguardar a mascara do modal sumir (CRITICAL FIX)
                WebDriverWait(driver, 5).until(EC.invisibility_of_element_located((By.CLASS_NAME, "noty_modal")))
                sleep(1)
            else:
                logging.info("Botão 'button-1' não encontrado desta vez.")
                
        except Exception as e:
            logging.info(f"Info: Modal search/close error: {e}")


        # Abrindo a planilha
        workbook = openpyxl.load_workbook(workbook_path)
        sheet = workbook["IMPRESSAO"]
        total_linhas = sheet.max_row

        for i in range(2, total_linhas + 1):
            
            num_guia = sheet.cell(row=i, column=1).value
            copias = sheet.cell(row=i, column=3).value or 1 

            if not num_guia:
                continue

            logging.info(f"Processando guia {num_guia} da linha Excel {i} com {copias} cópias")

            try:
                # Garantir que overlay sumiu antes de tudo
                try: WebDriverWait(driver, 2).until(EC.invisibility_of_element_located((By.CLASS_NAME, "noty_modal")))
                except: pass

                # Input Busca - Simplificado como no VBA
                try:
                    user_xpath = "/html/body/main/div[1]/div[1]/div[2]/div[1]/div[2]/input-text-search/div/div/div/input"
                    
                    try:
                        input_search = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, user_xpath)))
                    except:
                        # Fallback
                        alt_xpath = "/html/body/main/div[2]/div[1]/div[2]/div[1]/div[2]/input-text-search/div/div/div/input"
                        input_search = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, alt_xpath)))

                    input_search.clear()
                    input_search.send_keys(str(num_guia))
                    sleep(1)
                    
                    if input_search.get_attribute('value') != str(num_guia):
                        logging.warning("SendKeys falhou. Tentando JS fallback.")
                        driver.execute_script("arguments[0].value = arguments[1];", input_search, str(num_guia))

                except Exception as e:
                    logging.error(f"Erro input guia: {e}")
                    raise
                                                    /html/body/main/div[2]/div[1]/div[2]/div[1]/div[3]/input-periodo-data/div/div[1]/div/div/input
                # Limpar e interagir com campos de data e status
                try:
                    # Tentar via XPath relativo robusto primeiro (pega todos os inputs dentro da tag input-periodo-data)
                    date_inputs = driver.find_elements(By.XPATH, "//input-periodo-data//input")
                    
                    if len(date_inputs) >= 2:
                        # Limpa via JS para garantir
                        driver.execute_script("arguments[0].value = '';", date_inputs[0])
                        driver.execute_script("arguments[0].value = '';", date_inputs[1])
                        # Dispara eventos
                        driver.execute_script("arguments[0].dispatchEvent(new Event('change', { bubbles: true }));", date_inputs[0])
                        driver.execute_script("arguments[0].dispatchEvent(new Event('change', { bubbles: true }));", date_inputs[1])
                    else:
                        # Fallback para XPaths absolutos (ajustados para div[1])
                        try:
                            driver.find_element(By.XPATH, "/html/body/main/div[1]/div[1]/div[2]/div[1]/div[3]/input-periodo-data/div/div[1]/div/div/input").clear()
                            driver.find_element(By.XPATH, "/html/body/main/div[1]/div[1]/div[2]/div[1]/div[3]/input-periodo-data/div/div[2]/div/div/input").clear()
                        except:
                             # Tenta original (div[2])
                            driver.find_element(By.XPATH, "/html/body/main/div[2]/div[1]/div[2]/div[1]/div[3]/input-periodo-data/div/div[1]/div/div/input").clear()
                            driver.find_element(By.XPATH, "/html/body/main/div[2]/div[1]/div[2]/div[1]/div[3]/input-periodo-data/div/div[2]/div/div/input").clear()

                except Exception as e:
                    log_exception(e, "limpar campos de período", locator="XPATH=input-periodo-data")
                    # Não dar raise aqui para tentar seguir, o filtro de data pode não ser impeditivo se a guia for única
                    pass
                
                # Status (Robusto)
                try:
                    # Usando classe conhecida 'selectize-input' dentro de lista[2]
                    # Log anterior indicou: <div class="form-control selectize-input ...">
                    list_status = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, "//lista[2]//div[contains(@class, 'selectize-input')]"))
                    )
                    list_status.click()
                    sleep(0.5)
                    
                    # Selecionar "Todos" ou opção desejada (Assumindo que clica no dropdown e depois na opção)
                    # Se for apenas mudar o status, pode precisar de mais interações. 
                    # O código original apenas clicava na lista, vamos manter assim por enquanto.
                except Exception as e:
                    logging.info("Tentando fallback para Status...")
                    try:
                         # Fallback para estrutura div[1]
                         driver.find_element(By.XPATH, "/html/body/main/div[1]/div[1]/div[2]/div[1]/div[4]/lista[2]/div/div/div/div/div/div").click()
                    except:
                        log_exception(e, "abrir lista de status", locator="relative selectize-input")
                        raise
                
                sleep(1)

                # Localizar e imprimir guia
                try:
                    localizar_btn = wait_for(driver, By.ID, "localizar-procedimentos-btn", contexto="localizar procedimentos")
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", localizar_btn)
                    localizar_btn.click()
                except Exception as e:
                    log_exception(e, "clicar no botão Localizar", locator="By.ID=localizar-procedimentos-btn")
                    raise
                
                sleep(2)
                    
                # Ícone de Impressão (Robusto)
                try:
                    # Tentar encontrar pelo ícone de impressora (geralmente fa-print) ou link de impressão
                    # Estrutura genérica para item de grid
                    print_icon = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, "//a/i[contains(@class, 'fa-print')] | //a[contains(@title, 'Imprimir')]/i"))
                    )
                    # Scroll e Click
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", print_icon)
                    sleep(0.5)
                    print_icon.click()
                except Exception as e:
                     logging.info("Tentando fallback para Ícone de Impressão...")
                     try:
                        # Fallback div[1]
                        driver.find_element(By.XPATH, "/html/body/main/div[1]/div[2]/div/div[2]/div/div[2]/div[1]/div/div/div/div[1]/div/div/div[1]/a/i").click()
                     except:
                        log_exception(e, "clicar no ícone de impressão", locator="By.XPATH=//a/i[contains(@class, 'fa-print')]")
                        raise
                sleep(3)

                driver.switch_to.window(driver.window_handles[-1])

                # Ajustar zoom para 90% (pode tentar valores menores também)
                driver.execute_script("document.body.style.zoom='90%'")
                sleep(1)  # Aguarde um pouco para o ajuste de zoom

                # Configurar número de cópias para impressão
                for _ in range(copias):
                    try:
                        driver.execute_script("window.print();")
                    except Exception as e:
                        log_exception(e, "acionar impressão da guia")
                        raise
                    sleep(2)

                driver.close()
                driver.switch_to.window(driver.window_handles[1])
                sleep(10)

            except Exception as e:
                log_exception(e, f"processar guia {num_guia} (linha Excel {i})")
                try:
                    capture_screenshot(driver, f"guia_{num_guia}_linha_{i}")
                except Exception:
                    pass
                continue

    except Exception as e:
        logging.error(f"Erro durante a execução: {repr(e)}")
        logging.error("Trace:")
        logging.error("".join(traceback.format_exception(type(e), e, e.__traceback__)))
    finally:
        driver.quit()


# Executando a função
open_prestador()
