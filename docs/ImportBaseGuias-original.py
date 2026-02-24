import os
import time
import datetime
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from openpyxl import load_workbook, Workbook

# Variáveis globais
driver = None
tarefa = None
ARR_SADT = []
tempo = None
iterator = None
l_guia = None
continue_loop = False  # "continue" é palavra reservada, por isso renomeado
guiaAut = None
tafefa = None
cod_terminologia = None
saldoguiA = None
lin = None
cappen = False
colArr = None
arrterapias = [0] * 8  # Vetor para terapias
Benef_cart = None

# Classe auxiliar para imitar o objeto "By" do VBA
class ByWrapper:
    def ID(self, id_val):
        return (By.ID, id_val)
    def XPath(self, xpath_val):
        return (By.XPATH, xpath_val)
    def linktext(self, text):
        return (By.LINK_TEXT, text)

oCheck = ByWrapper()

# Função auxiliar para checar se um elemento está presente
def is_element_present(driver, by_locator, timeout=10):
    try:
        WebDriverWait(driver, timeout).until(EC.presence_of_element_located(by_locator))
        return True
    except TimeoutException:
        return False

# ======================================
# ROTINA: Remove_Hidden (habilita campos)
# ======================================
def Remove_Hidden(ElementRemove, driver):
    # Altera o atributo type de 'hidden' para 'text'
    driver.execute_script("arguments[0].setAttribute('type', 'text');", ElementRemove)

# ======================================
# ROTINA: funccarteira
# ======================================
def funccarteira(carteira, Retorno):
    """
    Divide a string 'carteira' assumindo o formato:
      p1.p2.p3-p4-p5
      
    Onde:
      p1: conteúdo antes do primeiro ponto.
      p2: conteúdo entre o primeiro e segundo ponto.
      p3: conteúdo entre o segundo e terceiro ponto.
      p4: conteúdo antes do primeiro traço após os pontos.
      p5: restante após o primeiro traço.
      
    Retorno:
      Se Retorno == 1, retorna p1;
      Se Retorno == 2, retorna p2;
      Se Retorno == 3, retorna p3;
      Se Retorno == 4, retorna p4;
      Se Retorno == 5, retorna p5.
    """
    try:
        # Divide nos três primeiros pontos
        p1, resto = carteira.split('.', 1)
        p2, resto = resto.split('.', 1)
        p3, resto = resto.split('.', 1)
        # Divide o restante com base no primeiro traço
        if '-' in resto:
            p4, p5 = resto.split('-', 1)
        else:
            p4, p5 = resto, ""
    except Exception as e:
        print("Erro na função funccarteira:", e)
        return ""
    
    # Remover espaços em branco extra
    p1 = p1.strip()
    p2 = p2.strip()
    p3 = p3.strip()
    p4 = p4.strip()
    p5 = p5.strip()
    
    if Retorno == 1:
        return p1
    elif Retorno == 2:
        return p2
    elif Retorno == 3:
        return p3
    elif Retorno == 4:
        return p4
    elif Retorno == 5:
        return p5
    else:
        return ""

# ======================================
# ROTINA: validCode
# ======================================
def validCode(driver, cod_terminologia):
    global arrterapias, colArr
    if cod_terminologia == "2250005103":
        if arrterapias[0] < 1500:
            colArr = 1
            return True
        else:
            return False
    elif cod_terminologia == "2250005111":
        if arrterapias[1] < 1500:
            colArr = 2
            return True
        else:
            return False
    elif cod_terminologia == "2250005189":
        if arrterapias[2] < 1500:
            colArr = 3
            return True
        else:
            return False
    elif cod_terminologia == "2250005170":
        if arrterapias[3] < 1500:
            colArr = 4
            return True
        else:
            return False
    elif cod_terminologia == "2250005278":
        if arrterapias[4] < 1500:
            colArr = 5
            return True
        else:
            return False
    elif cod_terminologia.startswith("50001213"):
        if arrterapias[5] < 1500:
            colArr = 6
            return True
        else:
            return False
    elif cod_terminologia.startswith("50000012"):
        if arrterapias[6] < 1500:
            colArr = 7
            return True
        else:
            return False
    return False

# ======================================
# ROTINA: importGuia
# ======================================
def importGuia(driver, lin):
    global Benef_cart
    try:
        wb_out = load_workbook("BaseGuiasImport2.xlsx")
        ws_out = wb_out.active
    except FileNotFoundError:
        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = "BD_Guias"
        headers = ["Carteirinha", "Paciente", "Guia", "Data_Autorização", "Senha", "Validade", "Código_Terapia", "Nome_Terapia", "Sessões Autorizadas"]
        ws_out.append(headers)
    MaxGuias = ws_out.max_row + 1

    cod_terminologia = ""
    saldoguiA = 0
    try:
        classguia_elements = driver.find_elements(By.CLASS_NAME, "MagnetoDataTD")
    except Exception as e:
        print("Erro ao obter elementos:", e)
        return
    countitem = 0
    cappen = False
    for element in classguia_elements:
        countitem += 1
        if countitem == 3:
            cod_terminologia = element.text[:10]
            cappen = validCode(driver, cod_terminologia)
        if countitem == 6:
            if cappen:
                try:
                    saldoguiA = int(element.text)
                except:
                    saldoguiA = 0
                time.sleep(1)
                if is_element_present(driver, oCheck.XPath('//*[@id="Button_Voltar"]')):
                    NewCarteira = driver.find_element(By.XPATH, '//*[@id="conteudo-submenu"]/form/table/tbody/tr[1]/td[2]')
                    textCarteira = NewCarteira.text
                    col1_value = textCarteira[:21]
                    col2_value = textCarteira[24:]
                    NewNumGuia = driver.find_element(By.XPATH, '//*[@id="conteudo-submenu"]/form/table/tbody/tr[3]/td[2]')
                    DataAuthorize = driver.find_element(By.XPATH, '//*[@id="conteudo-submenu"]/form/table/tbody/tr[4]/td[4]')
                    NewSenha = driver.find_element(By.XPATH, '//*[@id="conteudo-submenu"]/form/table/tbody/tr[5]/td[2]')
                    DataValid = driver.find_element(By.XPATH, '//*[@id="CampoValidadeSenha"]')
                    NewCodTerapia = driver.find_element(By.XPATH, '/html/body/div[1]/div[13]/div/table/tbody/tr[2]/td[3]/input')
                    QtdeSolicitado = driver.find_element(By.XPATH, '/html/body/div[1]/div[13]/div/table/tbody/tr[2]/td[5]')
                    QtdeAutorizado = driver.find_element(By.XPATH, '/html/body/div[1]/div[13]/div/table/tbody/tr[2]/td[6]')
                    ws_out.append([
                        col1_value,
                        col2_value,
                        NewNumGuia.text,
                        DataAuthorize.text,
                        NewSenha.text,
                        DataValid.text,
                        NewCodTerapia.get_attribute("value"),
                        QtdeSolicitado.text.strip(),
                        QtdeAutorizado.text.strip()
                    ])
                    driver.execute_script("window.scrollBy(0, 100);")
                    btnVoltar = driver.find_element(By.XPATH, '//*[@id="Button_Voltar"]')
                    btnVoltar.click()
                wb_out.save("BaseGuiasImport2.xlsx")
                return
            else:
                time.sleep(1)
                btnVoltar = driver.find_element(By.XPATH, '//*[@id="Button_Voltar"]')
                btnVoltar.click()
                return
    if countitem == 0:
        time.sleep(1)
        btnVoltar = driver.find_element(By.XPATH, '//*[@id="Button_Voltar"]')
        btnVoltar.click()
        time.sleep(1)
    wb_out.save("BaseGuiasImport2.xlsx")

# ======================================
# ROTINA: captura
# ======================================
def captura(driver):
    global Benef_cart, arrterapias
    x1 = funccarteira(Benef_cart, 1)
    x2 = funccarteira(Benef_cart, 2)
    x3 = funccarteira(Benef_cart, 3)
    x4 = funccarteira(Benef_cart, 4)
    x5 = funccarteira(Benef_cart, 5)
    
    if x1 != "0064":
        time.sleep(2)
        if is_element_present(driver, oCheck.XPath('//*[@id="Button_Consulta"]')):
            drvurl = driver.current_url
            driver.get(drvurl)
            time.sleep(5)
            consultabenef = driver.find_element(By.XPATH, '//*[@id="Button_Consulta"]')
            consultabenef.click()
            time.sleep(1)
            DT_VALIDADE_CARTAO = driver.find_element(By.XPATH, '//*[@id="DT_VALIDADE_CARTAO"]')
            DataValid = DT_VALIDADE_CARTAO.get_attribute("value")
            try:
                data_valid_date = datetime.datetime.strptime(DataValid, "%d/%m/%Y")
            except Exception as e:
                data_valid_date = datetime.datetime.now()
            if data_valid_date < datetime.datetime.now():
                x_date = (datetime.datetime.now() + datetime.timedelta(days=365)).strftime("%d/%m/%Y")
                DT_VALIDADE_CARTAO.click()
                time.sleep(1)
                driver.execute_script("document.getElementById('DT_VALIDADE_CARTAO').removeAttribute('readonly')")
                DT_VALIDADE_CARTAO.clear()
                DT_VALIDADE_CARTAO.send_keys(x_date)
            time.sleep(1)
            btn_atualiza = driver.find_element(By.XPATH, '//*[@id="Button_Update"]')
            btn_atualiza.click()
    
    countwait = 0
    while not is_element_present(driver, oCheck.XPath('//*[@id="s_NR_GUIA"]')):
        time.sleep(1)
        countwait += 1
        if countwait > 20:
            print("Erro de internet ou não foi liberado acesso às Guias do paciente")
            return

    cappen = False
    arrterapias = [0] * 8
    time.sleep(2)
    saldoguiA = 0
    
    try:
        if is_element_present(driver, oCheck.XPath('//*[@id="conteudo-submenu"]/table[2]/tbody/tr[1]/td[1]/a')):
            DataClassific = driver.find_element(By.XPATH, '//*[@id="conteudo-submenu"]/table[2]/tbody/tr[1]/td[1]/a')
            DataClassific.click()
            time.sleep(4)
            DataClassific = driver.find_element(By.XPATH, '//*[@id="conteudo-submenu"]/table[2]/tbody/tr[1]/td[1]/a')
            DataClassific.click()
        time.sleep(2)
        
        print("Iniciando processamento das guias...")
        while True:
            time.sleep(2)
            try:
                DataTable = driver.find_element(By.XPATH, '//*[@id="conteudo-submenu"]/table[2]')
                linhas = DataTable.find_elements(By.TAG_NAME, "tr")
                x_count = len(linhas)
                print(f"Encontradas {x_count} linhas na tabela")
                
                for idx in range(1, x_count - 1):
                    try:
                        print(f"Processando linha {idx + 1}")
                        # Rola até o elemento antes de interagir
                        elemento_span = driver.find_element(By.XPATH, f'//*[@id="conteudo-submenu"]/table[2]/tbody/tr[{idx+1}]/td[6]/span')
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", elemento_span)
                        time.sleep(1)
                        
                        if elemento_span.text == "Autorizado":
                            print("Guia Autorizada encontrada")
                            DataSolicit = driver.find_element(By.XPATH, f'//*[@id="conteudo-submenu"]/table[2]/tbody/tr[{idx+1}]/td[1]')
                            data_texto = DataSolicit.text.strip()
                            print(f"Data texto extraída: {data_texto}")
                            
                            try:
                                # Converte a data da guia considerando apenas dia/mês/ano
                                dataPEI = datetime.datetime.strptime(data_texto, "%d/%m/%Y").date()
                                print(f"Data da Guia convertida: {dataPEI.strftime('%d/%m/%Y')}")
                            except Exception as e:
                                print(f"Erro ao converter data '{data_texto}': {str(e)}")
                                dataPEI = datetime.datetime.now().date()
                            
                            # Calcula a data limite como data atual - 180 dias, sem horário
                            IniCompDate = datetime.datetime.now().date() - datetime.timedelta(days=270)
                            
                            print(f"Data da Guia: {dataPEI.strftime('%d/%m/%Y')}")
                            print(f"Data Limite: {IniCompDate.strftime('%d/%m/%Y')}")
                            
                            if dataPEI < IniCompDate:
                                print("Guia muito antiga (mais de 180 dias), parando processamento desta carteira...")
                                return  # Retorna para a função principal, parando o processamento desta carteira
                            
                            # Clica na guia para processamento
                            guia_link = driver.find_element(By.XPATH, f'//*[@id="conteudo-submenu"]/table[2]/tbody/tr[{idx+1}]/td[4]/a')
                            guia_link.click()
                            time.sleep(2)
                            importGuia(driver, idx+1)
                    except Exception as e:
                        print(f"Erro ao processar linha {idx + 1}: {str(e)}")
                        continue
                
                # Verifica se existe próxima página
                if is_element_present(driver, oCheck.linktext("Próxima")):
                    print("Mudando para próxima página...")
                    NUM_PAGE = driver.find_element(By.LINK_TEXT, "Próxima")
                    NUM_PAGE.click()
                    time.sleep(2)
                else:
                    print("Não há mais páginas para processar")
                    break
                    
            except Exception as e:
                print(f"Erro ao processar tabela: {str(e)}")
                break
                
    except Exception as e:
        print(f"Erro na função captura: {str(e)}")

# ======================================
# ROTINA: ConsultGuias – varre o arquivo "carteirinhas.xlsx"
# ======================================
def ConsultGuias(driver):
    global Benef_cart
    wb_in = load_workbook("carteirinhas.xlsx")
    ws_in = wb_in.active
    total_rows = ws_in.max_row
    
    print("Total de linhas a processar:", total_rows)
    
    for i in range(2, total_rows + 1):
        try:
            
            Benef_cart = ws_in.cell(row=i, column=1).value
            if not Benef_cart:
                print(f"Linha {i}: Carteira vazia, pulando...")
                continue     
                
            print(f"\nProcessando linha {i}, carteira: {Benef_cart}")
            
            # Extrai os valores da carteira usando a função funccarteira
            x1 = funccarteira(Benef_cart, 1)
            x2 = funccarteira(Benef_cart, 2)
            x3 = funccarteira(Benef_cart, 3)
            x4 = funccarteira(Benef_cart, 4)                                                                                                                
            x5 = funccarteira(Benef_cart, 5)
            
            # Imprime os valores para todas as carteiras
            print(f"Valores extraídos da carteira {Benef_cart}:")
            print("x1 =", x1)
            print("x2 =", x2)
            print("x3 =", x3)
            print("x4 =", x4)
            print("x5 =", x5)
            

            
            CountTry = 0
            while CountTry < 3:  # Limita a 3 tentativas
                try:
                    if is_element_present(driver, oCheck.XPath('//*[@id="cadastro_biometria"]/div/div[2]/span'), timeout=3):
                        print(f"Acessando visualização para carteira {Benef_cart}")
                        new_exame = driver.find_element(By.XPATH, '//*[@id="cadastro_biometria"]/div/div[2]/span')
                        new_exame.click()
                        time.sleep(2)  # Aumentei o tempo de espera após o clique
                        
                        # Troca para a nova janela
                        driver.switch_to.window(driver.window_handles[-1])
                        driver.maximize_window()
                        
                        
                        cartCompleto = x1 + x2 + x3 + x4 + x5      
                        cartaoParcial = x2 + x3 + x4 + x5
                            
                        print("Preenchendo campos para carteira:", Benef_cart)
                        
                        # Localiza e preenche os campos
                        element7 = driver.find_element(By.NAME, 'nr_via')
                        element6 = driver.find_element(By.NAME, 'DS_CARTAO')
                        element3 = driver.find_element(By.NAME, 'CD_DEPENDENCIA')
                        
                        # Habilita e preenche os campos
                        driver.execute_script("arguments[0].setAttribute('type', 'text');", element7)
                        time.sleep(1)
                        element7.send_keys(cartCompleto)
                        
                        driver.execute_script("arguments[0].setAttribute('type', 'text');", element6)
                        time.sleep(1)
                        element6.send_keys(cartaoParcial)
                        
                        driver.execute_script("arguments[0].setAttribute('type', 'text');", element3)
                        time.sleep(1)
                        element3.send_keys(x3)
                        
                        print("Campos preenchidos com sucesso")
                            
                        # Continua com o processamento...
                        captura(driver)
                        driver.close()
                        driver.switch_to.window(driver.window_handles[0])
                        break
                    else:
                        print(f"Tentativa {CountTry + 1}: Elemento não encontrado, atualizando página...")
                        driver.refresh()
                        time.sleep(2)
                        CountTry += 1
                except Exception as e:
                    print(f"Erro na tentativa {CountTry + 1}:", str(e))
                    CountTry += 1
                    if CountTry >= 3:
                        print(f"Falha após 3 tentativas para carteira {Benef_cart}")
                        driver.switch_to.window(driver.window_handles[0])
                        break
                    time.sleep(2)
        except Exception as e:
            print(f"Erro ao processar carteira {Benef_cart}:", str(e))
            continue
    
    print("\nProcessamento finalizado")
    driver.quit()

# ======================================
# ROTINA: SGUCARD (ponto de entrada)
# ======================================
def SGUCARD():
    global driver
    chrome_options = Options()
    username = os.environ.get("USERNAME") or os.environ.get("USER")
    #ChromePath = f"C:\\Users\\{username}\\AppData\\Local\\Google\\Chrome\\User Data"
    #chrome_options.add_argument(f"--user-data-dir={ChromePath}")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-infobars")
    # Para rodar em modo headless descomente a linha a seguir:
    # chrome_options.add_argument("--headless")
    driver = webdriver.Chrome(options=chrome_options)
    
    time.sleep(3)
    driver.get("https://sgucard.unimedgoiania.coop.br/cmagnet/Login.do")
    driver.maximize_window()
    
    while not is_element_present(driver, oCheck.ID("passwordTemp"), timeout=3):
        time.sleep(1)
        
    login_elem = driver.find_element(By.ID, "login")
    passwordTemp = driver.find_element(By.ID, "passwordTemp")
    Button_DoLogin = driver.find_element(By.ID, "Button_DoLogin")
    login_elem.send_keys("REC2209525")
    time.sleep(1)
    passwordTemp.clear()
    passwordTemp.send_keys("Unimed@2025")
    Button_DoLogin.click()
    time.sleep(4)
    ConsultGuias(driver)

if __name__ == "__main__":
    SGUCARD()